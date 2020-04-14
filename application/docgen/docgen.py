# -*- coding: utf-8 -*-
import datetime

import os
from application.models.inventory import Inventory
from application.nutils.excel import header_value_tuple
from flask import current_app
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font

TEMPLATE_FOLDER = os.path.join(current_app.root_path, 'docgen', 'templates')
OUTPUT_FOLDER = os.path.join(current_app.root_path, 'docgen', 'output')

if not os.path.exists(OUTPUT_FOLDER):
    os.makedirs(OUTPUT_FOLDER)


def generate_ops_overview_excel(start, end, stores_data):
    wb = load_workbook(filename=os.path.join(TEMPLATE_FOLDER, 'ops_overview_report.xlsx'))

    ws = wb.worksheets[0]

    DATA_ROW_START = 3
    DATA_COLUMN_START = 1
    for index, store_data in enumerate(stores_data):
        ws.cell(row=index + DATA_ROW_START, column=DATA_COLUMN_START, value='%s' % store_data[1])
        ws.cell(row=index + DATA_ROW_START, column=DATA_COLUMN_START + 1, value=store_data[2]['online_days'])
        ws.cell(row=index + DATA_ROW_START, column=DATA_COLUMN_START + 2, value=store_data[2]['rx_count'])
        ws.cell(row=index + DATA_ROW_START, column=DATA_COLUMN_START + 3, value=store_data[2]['incomplete_rx_count'])
        per_cell = ws.cell(row=index + DATA_ROW_START, column=DATA_COLUMN_START + 4,
                           value=store_data[2]['incomplete_rx_percent'])
        per_cell.alignment = per_cell.alignment.copy(horizontal='right', vertical='bottom')
        ws.cell(row=index + DATA_ROW_START, column=DATA_COLUMN_START + 5,
                value=store_data[2]['rx_customer_count_in_draft'])
        ws.cell(row=index + DATA_ROW_START, column=DATA_COLUMN_START + 6, value=store_data[2]['new_orders_count'])
        ws.cell(row=index + DATA_ROW_START, column=DATA_COLUMN_START + 7,
                value=store_data[2]['no_deliver_date_orders_count'])
        ws.cell(row=index + DATA_ROW_START, column=DATA_COLUMN_START + 8, value=store_data[2]['overdue_orders_count'])
        ws.cell(row=index + DATA_ROW_START, column=DATA_COLUMN_START + 9,
                value=store_data[2]['imported_customer_count'])
        ws.cell(row=index + DATA_ROW_START, column=DATA_COLUMN_START + 10, value=store_data[2]['new_calls_count'])
        ws.cell(row=index + DATA_ROW_START, column=DATA_COLUMN_START + 11, value=store_data[2]['new_appt_count'])

    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    STANDARD_FONT = Font(name='Microsoft YaHei', size=11)
    TITLE_FONT = Font(name='Microsoft YaHei', size=12)
    for row in ws.rows:
        for cell in row:
            cell.border = thin_border
            if cell.row <= 2:
                cell.font = TITLE_FONT
            else:
                cell.font = STANDARD_FONT

    ws.cell(row=1, column=1, value='汽车经销商集团u客系列使用监测结果(%s～%s)' % (start, end))

    filename = 'ops_overview_report_%s.xlsx' % datetime.datetime.now().strftime('%Y%m%d%H%M%S')
    wb.save(filename=os.path.join(OUTPUT_FOLDER, filename))

    return filename


def generate_store_inventories_excel(store_id, inventories):
    wb = load_workbook(filename=os.path.join(TEMPLATE_FOLDER, 'inv_template.xlsx'))
    ws = wb.worksheets[0]

    DATA_ROW_START = 2
    DATA_COLUMN_START = 1
    col_names = header_value_tuple(ws.rows[0])
    for index, data in enumerate(inventories):
        for col_index, col_name in enumerate(col_names):
            ws.cell(row=DATA_ROW_START + index, column=DATA_COLUMN_START + col_index,
                    value=getattr(data, Inventory.fields_mapper.get(col_name, ''), ''))

    add_border(ws)
    filename = 'store_inventories_%s.xlsx' % datetime.datetime.now().strftime('%Y%m%d%H%M%S')
    wb.save(filename=os.path.join(OUTPUT_FOLDER, filename))

    return filename


def add_border(ws):
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    for row in ws.rows:
        for cell in row:
            cell.border = thin_border
