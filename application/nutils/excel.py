# -*- coding: utf-8 -*-
from collections import OrderedDict

from openpyxl import load_workbook
import re

descriptions = {'intent-car': u'车型', 'test-drive-car': u'试驾车型', 'intent-color': u'颜色', 'intent-level': u'意向周期',
                'channel': u'来源渠道'}


def safe_convert_unicode(val):
    if isinstance(val, int):
        return val
    if isinstance(val, long):
        return unicode(val)
    val = val or ''
    return unicode(val.strip())


def convert_value_to_code(val):
    if val == u"无":
        return 'none'
    val = re.sub(r'\s+', '_', val)
    return val


def header_value_tuple(header):
    ret = []
    for cell in header:
        if cell.value and len(unicode(cell.value).strip()) > 0:
            ret.append(cell.value)
    return tuple(ret)


def row_value_tuple(row, cols_num=None):
    if not cols_num:
        cols_num = len(row)

    ret = []
    for index, cell in enumerate(row):
        if index < cols_num:
            ret.append(cell.value)

    return tuple(ret)


def parse_excel(filename):
    wb = load_workbook(filename=filename)

    try:
        ws = wb.worksheets[0]

        headers = header_value_tuple(ws.rows[0])

        data = []
        cols_num = len(headers)
        for index, row in enumerate(ws.rows):
            if index > 0:
                data.append(row_value_tuple(row, cols_num))

        return [OrderedDict(zip(headers, data)) for data in data]
    except Exception:
        raise Exception('errors happen while parsing excel %s' % filename)


def _get_sheet_data(filename):
    wb = load_workbook(filename=filename, read_only=True)

    lookupvalues_data = dict()
    lookup_list = []

    for name in descriptions:
        description = descriptions[name]
        ws = wb.get_sheet_by_name(description)
        row_lookup = ws.rows
        for row in row_lookup:
            line = [safe_convert_unicode(col.value) for col in row]

            if not line[0].startswith("#"):
                if name in ['intent-car', 'test-drive-car', 'intent-color']:
                    count = 2
                    code = convert_value_to_code(line[0])
                    line.insert(0, code)
                    count += 1
                    if len(line) == count:
                        lookup_list.append(line)
                    else:
                        return False
                else:
                    count = 3
                    if len(line) == count:
                        lookup_list.append(line)
                    else:
                        return False

        lookupvalues_data[name] = lookup_list
        lookup_list = []

    return lookupvalues_data


def load_excel(filename):
    try:
        return _get_sheet_data(filename)
    except Exception:
        raise Exception('errors happen while parsing lookupvalues excel %s' % filename)
